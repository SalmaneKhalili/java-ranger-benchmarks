#!/usr/bin/env python3
"""Extract JRBench results from XML + logs into an Excel workbook.

Single mode (1 dir):
  Sheet 1: Overview — per-suite stats (total, correct, incorrect, unknown, correct rate).
  Sheet 2: Benchmarks — every benchmark with walltime and full error traces.

Comparison mode (2 dirs):
  Sheet 1: Overview — side-by-side per-suite comparison (NOFP vs FP).
  Sheet 2: Benchmarks — one row per benchmark with side-by-side NOFP/FP
             columns (status, walltime, error) + Status Change column.

Usage:
  extract-results.py <results-dir> [<results-dir-2>] [-o FILE.xlsx]
"""

import argparse, bz2, glob, os, re, sys
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_results_xml(fpath):
    """Yield dicts: {suite, name, status, expected, walltime, error, logfile}."""
    try:
        if fpath.endswith(".bz2"):
            with bz2.open(fpath) as f:
                tree = ET.parse(f)
        else:
            tree = ET.parse(fpath)
    except Exception as e:
        print(f"  Warning: could not parse {fpath}: {e}", file=sys.stderr)
        return

    root = tree.getroot()
    for run in root.findall("run"):
        suite = run.get("suite") or root.get("benchmarkname", "unknown")
        name = (run.get("name") or "").rsplit("/", 1)[-1]
        name = name.replace(".yml", "").replace(".java", "")
        expected = run.get("expectedVerdict", "?")

        cols = {}
        for col in run.findall("column"):
            cols[col.get("title")] = col.get("value")

        status = cols.get("status", "unknown")
        error = ""
        error_el = run.find("column[@title='error']")
        if error_el is not None:
            error = error_el.get("value", "")
        if not error and status != "correct":
            category_el = run.find("column[@title='category']")
            if category_el is not None and category_el.get("value") == "error":
                error = cols.get("error", "")

        raw_wt = cols.get("walltime", "0s")
        walltime = 0.0
        m = re.search(r"([\d.]+)", raw_wt)
        if m:
            walltime = float(m.group(1))

        yield {
            "suite": suite,
            "name": name,
            "status": status,
            "expected": expected,
            "walltime": walltime,
            "error": error,
            "logfile": run.get("logfile", ""),
        }


def extract_full_error(log_path):
    """Extract the first error block + tail from a JPF .log file."""
    if not log_path or not os.path.isfile(log_path):
        return ""
    try:
        with open(log_path, errors="replace") as f:
            text = f.read()
    except OSError:
        return ""

    lines = text.splitlines()
    stripped = [l.rstrip() for l in lines]

    # Find the first line that signals an error
    start = -1
    for i, s in enumerate(stripped):
        low = s.lower()
        if "[severe]" in low:
            start = i
            break
        if "exception" in low:
            start = i
            break
        if low.startswith("error #") or low.startswith("error:"):
            start = i
            break

    if start >= 0:
        return "\n".join(stripped[start:])

    # Check the results section for error lines (skip "no errors detected")
    for i, s in enumerate(stripped):
        if "===== results" in s or "========== results" in s:
            tail = "\n".join(stripped[i:])
            if "error" in tail.lower() and "no errors detected" not in tail.lower():
                return tail
            break

    return ""


def safe_str(v, maxlen=32767):
    s = str(v) if v is not None else ""
    return s[:maxlen]


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _suite_stats(records):
    suites = {}
    for rec in records:
        s = rec["suite"]
        if s not in suites:
            suites[s] = {"total": 0, "correct": 0, "incorrect": 0, "unknown": 0}
        suites[s]["total"] += 1
        st = rec["status"]
        if st == "correct":
            suites[s]["correct"] += 1
        elif st == "incorrect":
            suites[s]["incorrect"] += 1
        else:
            suites[s]["unknown"] += 1
    return suites


def _rate_fill(rate):
    if rate >= 90:
        return GOOD_FILL
    elif rate >= 50:
        return WARN_FILL
    else:
        return BAD_FILL


def build_overview_sheet(wb, records):
    ws = wb.active
    ws.title = "Overview"

    headers = ["Suite", "Total", "Correct", "Incorrect", "Unknown", "Correct Rate"]
    ws.append(headers)
    style_header(ws, len(headers))

    ws.column_dimensions["A"].width = 30
    for c in ["B", "C", "D", "E"]:
        ws.column_dimensions[c].width = 12
    ws.column_dimensions["F"].width = 14

    suites = _suite_stats(records)
    totals = {"total": 0, "correct": 0, "incorrect": 0, "unknown": 0}

    for sname in sorted(suites):
        st = suites[sname]
        rate = st["correct"] / st["total"] * 100 if st["total"] else 0
        ws.append([sname, st["total"], st["correct"], st["incorrect"], st["unknown"],
                    f"{rate:.1f}%"])
        r = ws.max_row
        cell = ws.cell(row=r, column=6)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = _rate_fill(rate)
        for k in totals:
            totals[k] += st[k]

    ws.append([])
    r_total = ws.max_row + 1
    total_rate = totals["correct"] / totals["total"] * 100 if totals["total"] else 0
    row_data = ["TOTAL", totals["total"], totals["correct"], totals["incorrect"],
                totals["unknown"], f"{total_rate:.1f}%"]
    ws.append(row_data)
    for col in range(1, len(row_data) + 1):
        ws.cell(row=r_total, column=col).font = Font(bold=True, size=12)
    rate_cell = ws.cell(row=r_total, column=6)
    rate_cell.alignment = Alignment(horizontal="center")
    rate_cell.fill = _rate_fill(total_rate)

    return ws


def build_comparison_overview_sheet(wb, records1, records2, label1, label2):
    ws = wb.active
    ws.title = "Overview"

    headers = ["Suite", "Total",
               f"Correct {label1}", f"Incorrect {label1}", f"Unknown {label1}", f"Rate {label1}",
               f"Correct {label2}", f"Incorrect {label2}", f"Unknown {label2}", f"Rate {label2}",
               "Δ Rate"]
    ws.append(headers)
    style_header(ws, len(headers))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 8
    for c in ["C", "D", "E", "G", "H", "I"]:
        ws.column_dimensions[c].width = 14
    for c in ["F", "J"]:
        ws.column_dimensions[c].width = 12
    ws.column_dimensions["K"].width = 10

    s1 = _suite_stats(records1)
    s2 = _suite_stats(records2)
    all_suites = sorted(set(list(s1.keys()) + list(s2.keys())))

    totals = {"total": 0,
              label1: {"total": 0, "correct": 0, "incorrect": 0, "unknown": 0},
              label2: {"total": 0, "correct": 0, "incorrect": 0, "unknown": 0}}

    for sname in all_suites:
        st1 = s1.get(sname, {"total": 0, "correct": 0, "incorrect": 0, "unknown": 0})
        st2 = s2.get(sname, {"total": 0, "correct": 0, "incorrect": 0, "unknown": 0})
        t = st1["total"]
        rate1 = st1["correct"] / t * 100 if t else 0
        rate2 = st2["correct"] / t * 100 if t else 0
        delta = rate2 - rate1
        ws.append([sname, t,
                    st1["correct"], st1["incorrect"], st1["unknown"], f"{rate1:.1f}%",
                    st2["correct"], st2["incorrect"], st2["unknown"], f"{rate2:.1f}%",
                    f"{delta:+.1f}%"])
        r = ws.max_row
        for c in [6, 10]:
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=c).fill = _rate_fill(rate1 if c == 6 else rate2)
        dc = ws.cell(row=r, column=11)
        dc.alignment = Alignment(horizontal="center")
        if delta > 0:
            dc.fill = GOOD_FILL
        elif delta < 0:
            dc.fill = BAD_FILL

        totals["total"] += t
        for k in totals[label1]:
            totals[label1][k] += st1[k]
        for k in totals[label2]:
            totals[label2][k] += st2[k]

    ws.append([])
    r_total = ws.max_row + 1
    tt = totals["total"]
    rate1 = totals[label1]["correct"] / tt * 100 if tt else 0
    rate2 = totals[label2]["correct"] / tt * 100 if tt else 0
    delta = rate2 - rate1
    row_data = ["TOTAL", tt,
                totals[label1]["correct"], totals[label1]["incorrect"],
                totals[label1]["unknown"], f"{rate1:.1f}%",
                totals[label2]["correct"], totals[label2]["incorrect"],
                totals[label2]["unknown"], f"{rate2:.1f}%",
                f"{delta:+.1f}%"]
    ws.append(row_data)
    for col in range(1, len(row_data) + 1):
        ws.cell(row=r_total, column=col).font = Font(bold=True, size=12)
    for c in [6, 10]:
        ws.cell(row=r_total, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=r_total, column=c).fill = _rate_fill(rate1 if c == 6 else rate2)
    dc = ws.cell(row=r_total, column=11)
    dc.alignment = Alignment(horizontal="center")
    if delta > 0:
        dc.fill = GOOD_FILL
    elif delta < 0:
        dc.fill = BAD_FILL

    return ws


def build_benchmarks_sheet(wb, records, results_dir, mode_label=None):
    ws = wb.create_sheet("Benchmarks")

    if mode_label:
        headers = ["Mode", "Suite", "Benchmark", "Expected", "Status", "Walltime (s)", "Error Message"]
    else:
        headers = ["Suite", "Benchmark", "Expected", "Status", "Walltime (s)", "Error Message"]
    ws.append(headers)
    style_header(ws, len(headers))

    ws.column_dimensions["A"].width = 8 if mode_label else 28
    ws.column_dimensions["B"].width = 28 if mode_label else 42
    ws.column_dimensions["C"].width = 42 if mode_label else 12
    ws.column_dimensions["D"].width = 12 if mode_label else 14
    ws.column_dimensions["E"].width = 14 if mode_label else 14
    ws.column_dimensions["F"].width = 14 if mode_label else 100
    if mode_label:
        ws.column_dimensions["G"].width = 100

    for rec in records:
        log_path = os.path.join(results_dir, "logs", f"{rec['name']}.log")

        error_msg = ""
        if rec["status"] != "correct":
            full_error = extract_full_error(log_path)
            error_msg = full_error or rec["error"]

        if mode_label:
            row = [
                mode_label,
                safe_str(rec["suite"]),
                safe_str(rec["name"]),
                safe_str(rec["expected"]),
                safe_str(rec["status"]),
                rec["walltime"],
                safe_str(error_msg),
            ]
        else:
            row = [
                safe_str(rec["suite"]),
                safe_str(rec["name"]),
                safe_str(rec["expected"]),
                safe_str(rec["status"]),
                rec["walltime"],
                safe_str(error_msg),
            ]

        ws.append(row)

        r = ws.max_row
        wt_col = 6 if mode_label else 5
        wt_cell = ws.cell(row=r, column=wt_col)
        wt_cell.number_format = "0.00"

        st_col = 5 if mode_label else 4
        status_cell = ws.cell(row=r, column=st_col)
        if rec["status"] == "correct":
            status_cell.fill = GOOD_FILL
        elif rec["status"] == "incorrect":
            status_cell.fill = BAD_FILL
        else:
            status_cell.fill = WARN_FILL

    return ws


def _load_dir(results_dir):
    if not os.path.isdir(results_dir):
        print(f"Error: {results_dir} is not a directory", file=sys.stderr)
        sys.exit(1)

    xml_files = sorted(glob.glob(os.path.join(results_dir, "**", "*.results.xml*"),
                                 recursive=True))
    if not xml_files:
        xml_files = sorted(glob.glob(os.path.join(results_dir, "*.results.xml*")))

    if not xml_files:
        print(f"No .results.xml files found in {results_dir}", file=sys.stderr)
        sys.exit(1)

    records = []
    for fpath in xml_files:
        for rec in parse_results_xml(fpath):
            records.append(rec)

    if not records:
        print(f"No benchmark records found in {results_dir}", file=sys.stderr)
        sys.exit(1)

    return records


def main():
    parser = argparse.ArgumentParser(
        description="Extract benchmark results to Excel (overview + benchmarks sheets)"
    )
    parser.add_argument("results_dirs", nargs="+",
                        help="1 or 2 directories containing result XMLs and logs/")
    parser.add_argument("--output", "-o", default="results.xlsx",
                        help="Output Excel file path (default: results.xlsx)")
    args = parser.parse_args()

    if len(args.results_dirs) == 1:
        records = _load_dir(args.results_dirs[0])
        wb = Workbook()
        build_overview_sheet(wb, records)
        build_benchmarks_sheet(wb, records, args.results_dirs[0])
        out_path = args.output
        wb.save(out_path)
        n_correct = sum(1 for r in records if r["status"] == "correct")
        print(f"Wrote {len(records)} benchmarks ({n_correct} correct, "
              f"{len(records) - n_correct} others) to {out_path}")

    elif len(args.results_dirs) == 2:
        d1, d2 = args.results_dirs
        records1 = _load_dir(d1)
        records2 = _load_dir(d2)
        label1, label2 = "NOFP", "FP"
        wb = Workbook()
        build_comparison_overview_sheet(wb, records1, records2, label1, label2)
        ws = wb.create_sheet("Benchmarks")

        headers = [
            "Suite", "Benchmark", "Expected",
            f"Status {label1}", f"Walltime {label1} (s)", f"Error {label1}",
            f"Status {label2}", f"Walltime {label2} (s)", f"Error {label2}",
            "Status Change",
        ]
        ws.append(headers)
        style_header(ws, len(headers))
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 50
        ws.column_dimensions["J"].width = 18

        def _resolve_log(rec, results_dir):
            return os.path.join(results_dir, "logs", f"{rec['name']}.log")

        def _get_error(rec, results_dir):
            if rec["status"] == "correct":
                return ""
            full = extract_full_error(_resolve_log(rec, results_dir))
            return full or rec["error"]

        idx1 = {(r["suite"], r["name"]): r for r in records1}
        idx2 = {(r["suite"], r["name"]): r for r in records2}
        all_keys = sorted(set(idx1.keys()) | set(idx2.keys()))

        CHANGE_FILLS = {
            "improved": GOOD_FILL,
            "regressed": BAD_FILL,
            "sidegrade": WARN_FILL,
        }

        improved = regressed = sidegrade = same = 0

        for key in all_keys:
            r1 = idx1.get(key)
            r2 = idx2.get(key)
            suite, name = key
            expected = r1["expected"] if r1 else r2["expected"]

            s1 = r1["status"] if r1 else "—"
            s2 = r2["status"] if r2 else "—"
            wt1 = r1["walltime"] if r1 else 0.0
            wt2 = r2["walltime"] if r2 else 0.0
            err1 = _get_error(r1, d1) if r1 else ""
            err2 = _get_error(r2, d2) if r2 else ""

            if s1 == s2 or s1 == "—" or s2 == "—":
                change = ""
                change_type = "same"
            elif s2 == "correct":
                change = f"{s1}→{s2}"
                change_type = "improved"
            elif s1 == "correct":
                change = f"{s1}→{s2}"
                change_type = "regressed"
            else:
                change = f"{s1}→{s2}"
                change_type = "sidegrade"

            ws.append([
                safe_str(suite), safe_str(name), safe_str(expected),
                safe_str(s1), wt1, safe_str(err1),
                safe_str(s2), wt2, safe_str(err2),
                safe_str(change),
            ])

            r = ws.max_row
            for c in [5, 8]:
                ws.cell(row=r, column=c).number_format = "0.00"
            for c in [4, 7]:
                val = ws.cell(row=r, column=c).value
                if val == "correct":
                    ws.cell(row=r, column=c).fill = GOOD_FILL
                elif val == "incorrect":
                    ws.cell(row=r, column=c).fill = BAD_FILL
                elif val == "unknown":
                    ws.cell(row=r, column=c).fill = WARN_FILL

            if change_type != "same":
                ws.cell(row=r, column=10).fill = CHANGE_FILLS[change_type]

            if change_type == "improved":
                improved += 1
            elif change_type == "regressed":
                regressed += 1
            elif change_type == "sidegrade":
                sidegrade += 1
            else:
                same += 1

        ws.append([])
        r_summary = ws.max_row + 1
        ws.append([
            "SUMMARY", "", "",
            "", "", "",
            "", "", "",
            f"same={same} improved={improved} regressed={regressed} sidegrade={sidegrade}",
        ])
        for col in range(1, 11):
            ws.cell(row=r_summary, column=col).font = Font(bold=True)
        ws.cell(row=r_summary + 1, column=10).font = Font(bold=True)

        out_path = args.output
        wb.save(out_path)
        n1 = sum(1 for r in records1 if r["status"] == "correct")
        n2 = sum(1 for r in records2 if r["status"] == "correct")
        print(f"Wrote {len(all_keys)} benchmarks side-by-side ({same} same, "
              f"{improved} improved, {regressed} regressed, {sidegrade} sidegrade) "
              f"to {out_path}")

    else:
        print("Error: provide 1 (single mode) or 2 (comparison mode) results directories",
              file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
